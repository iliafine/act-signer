"""
Батч-режим: подписание всех актов из реестра, у которых в столбце
«Принятие ИД» указано «Принято».

Вход: zip-архив, в котором лежит:
  - файл реестра (.xlsx) со столбцом «Принятие ИД» и столбцом-идентификатором
    папки акта (номер/дата),
  - папки актов (например «2-от»), внутри каждой — файл «Акт*.docx».

ВНИМАНИЕ: точное название столбца с идентификатором папки и точный формат
имени папки (например «2-от» против «2-о») пока не подтверждены реальным
файлом реестра — сопоставление сделано через нечёткое совпадение (rapidfuzz)
с порогом FOLDER_MATCH_THRESHOLD и подробным логом. После получения реального
zip нужно прогнать и при необходимости скорректировать STATUS_COLUMN /
FOLDER_COLUMN / порог совпадения.

Запуск:
    python batch_sign.py --input реестр_и_акты.zip --output подписанные.zip
"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
import tempfile
import zipfile
from pathlib import Path

import openpyxl
from rapidfuzz import fuzz

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))
from app.docx_signer import sign_document  # noqa: E402

STATUS_COLUMN_CANDIDATES = ["принятие ид", "принятие ид."]
STATUS_VALUE = "принято"
FOLDER_COLUMN_CANDIDATES = ["номер", "№", "номер и дата", "документ"]
FOLDER_MATCH_THRESHOLD = 70
ACT_FILE_PREFIX = "акт"


def _normalize(text: str) -> str:
    return " ".join(str(text or "").strip().lower().split())


def find_registry_xlsx(root: Path) -> Path:
    candidates = list(root.rglob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError("В zip не найден файл реестра (.xlsx)")
    return candidates[0]


def read_registry_rows(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb.active

    header_row_idx = None
    header_map: dict[str, int] = {}
    for row_idx in range(1, min(sheet.max_row, 10) + 1):
        values = [_normalize(c.value) for c in sheet[row_idx]]
        if any(cand in v for v in values for cand in STATUS_COLUMN_CANDIDATES):
            header_row_idx = row_idx
            for col_idx, v in enumerate(values, start=1):
                if v:
                    header_map[v] = col_idx
            break

    if header_row_idx is None:
        raise ValueError(
            "Не нашёл строку заголовка со столбцом «Принятие ИД». "
            "Нужно скорректировать STATUS_COLUMN_CANDIDATES под реальный реестр."
        )

    status_col = next((idx for name, idx in header_map.items() if any(c in name for c in STATUS_COLUMN_CANDIDATES)), None)
    folder_col = next((idx for name, idx in header_map.items() if any(c in name for c in FOLDER_COLUMN_CANDIDATES)), None)

    if status_col is None:
        raise ValueError("Не нашёл колонку статуса «Принятие ИД»")
    if folder_col is None:
        raise ValueError(
            "Не нашёл колонку с номером/идентификатором папки акта. "
            "Уточните FOLDER_COLUMN_CANDIDATES под реальный реестр."
        )

    rows = []
    for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
        status_val = _normalize(sheet.cell(row_idx, status_col).value)
        folder_val = sheet.cell(row_idx, folder_col).value
        if not folder_val:
            continue
        rows.append({"row": row_idx, "status": status_val, "folder_id": str(folder_val).strip()})
    return rows


def find_matching_folder(root: Path, folder_id: str) -> Path | None:
    target = _normalize(folder_id).replace("/", "-").replace(".", "-").replace(" ", "-")
    best: tuple[Path | None, float] = (None, 0)
    for candidate in root.iterdir():
        if not candidate.is_dir():
            continue
        cand_norm = _normalize(candidate.name).replace("/", "-").replace(".", "-").replace(" ", "-")
        score = fuzz.ratio(target, cand_norm)
        if score > best[1]:
            best = (candidate, score)
    if best[1] >= FOLDER_MATCH_THRESHOLD:
        return best[0]
    return None


def find_act_file(folder: Path) -> Path | None:
    for f in folder.rglob("*"):
        if f.is_file() and f.suffix.lower() == ".docx" and _normalize(f.stem).startswith(ACT_FILE_PREFIX):
            return f
    return None


def run(input_zip: Path, output_zip: Path) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        extract_dir = tmp_path / "extracted"
        extract_dir.mkdir()
        with zipfile.ZipFile(input_zip) as zf:
            zf.extractall(extract_dir)

        registry_path = find_registry_xlsx(extract_dir)
        rows = read_registry_rows(registry_path)

        output_dir = tmp_path / "output"
        output_dir.mkdir()
        report: list[dict] = []

        for row in rows:
            entry = {"row": row["row"], "folder_id": row["folder_id"], "status": row["status"]}

            if STATUS_VALUE not in row["status"]:
                entry["result"] = "skipped_not_accepted"
                report.append(entry)
                continue

            folder = find_matching_folder(extract_dir, row["folder_id"])
            if folder is None:
                entry["result"] = "folder_not_found"
                report.append(entry)
                continue
            entry["matched_folder"] = folder.name

            act_file = find_act_file(folder)
            if act_file is None:
                entry["result"] = "act_file_not_found"
                report.append(entry)
                continue
            entry["act_file"] = act_file.name

            out_file = output_dir / folder.name / act_file.name
            try:
                sign_report = sign_document(str(act_file), str(out_file))
                entry["result"] = "signed" if sign_report.all_signed else "signed_with_issues"
                entry["sign_details"] = [
                    {"role": r.role_text, "status": r.status, "signer": r.matched_signer.full_name if r.matched_signer else None}
                    for r in sign_report.results
                ]
            except Exception as e:  # noqa: BLE001
                entry["result"] = f"error: {e}"

            report.append(entry)

        (output_dir / "_batch_report.json").write_text(
            json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
        )

        output_zip.parent.mkdir(parents=True, exist_ok=True)
        if output_zip.exists():
            output_zip.unlink()
        with zipfile.ZipFile(output_zip, "w", zipfile.ZIP_DEFLATED) as zf:
            for item in output_dir.rglob("*"):
                if item.is_file():
                    zf.write(item, item.relative_to(output_dir))

        signed = sum(1 for r in report if r.get("result") == "signed")
        issues = sum(1 for r in report if r.get("result") not in ("signed", "skipped_not_accepted"))
        print(f"Готово: {signed} подписано без замечаний, {issues} с проблемами/пропущено. Отчёт: _batch_report.json внутри {output_zip}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Батч-подписание актов по реестру")
    parser.add_argument("--input", required=True, type=Path, help="zip с реестром и папками актов")
    parser.add_argument("--output", required=True, type=Path, help="путь к выходному zip с подписанными актами")
    args = parser.parse_args()
    run(args.input, args.output)
